import openpyxl
from openpyxl.chart import BarChart, Reference
from datetime import datetime


class ExcelExporter:
    def __init__(self, include_video_info=True, include_creator_info=True):
        self.include_video_info = include_video_info
        self.include_creator_info = include_creator_info

    # =======================================================
    # Export to Excel
    # =======================================================
    def export(self, data):
        filename = f"yt_crawler_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "YouTube Results"

        # ======================
        # Define Columns
        # ======================
        columns = []
        if self.include_video_info:
            columns.extend([
                ("keyword", "Keyword"),
                ("video_url", "Video URL"),
                ("title", "Title"),
                ("description", "Description"),
                ("views", "Views"),
                ("likes", "Likes"),
                ("comments", "Comments"),
                ("publishedAt", "Published At")
            ])

        if self.include_creator_info:
            columns.extend([
                ("channel_title", "Channel Name"),
                ("channel_url", "Channel URL"),
                ("channel_country", "Country"),
                ("subscribers", "Subscribers")
            ])

        # Write header
        for col, (_, header) in enumerate(columns, start=1):
            ws.cell(row=1, column=col, value=header)

        # ======================
        # Write rows
        # ======================
        for row_index, item in enumerate(data, start=2):
            for col_index, (key, _) in enumerate(columns, start=1):
                ws.cell(row=row_index, column=col_index, value=item.get(key, ""))

        # ======================
        # Add Bar Chart (Views)
        # ======================
        try:
            views_col_index = [i for i, (key, _) in enumerate(columns, start=1) if key == "views"]
            if views_col_index:
                col = views_col_index[0]
                chart = BarChart()
                chart.title = "Video Views Chart"
                chart.y_axis.title = "Views"
                chart.x_axis.title = "Video"

                data_ref = Reference(ws, min_col=col, max_col=col, min_row=1, max_row=len(data) + 1)
                cats_ref = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=len(data) + 1)  # video title

                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)

                ws.add_chart(chart, f"A{len(data) + 3}")
        except Exception as e:
            print("Chart creation failed:", e)

        # Save
        wb.save(filename)
        return filename
